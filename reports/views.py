from pyexpat.errors import messages
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, F
from django.utils import timezone
from datetime import timedelta, datetime
from django.http import HttpResponse
import csv
from orders.models import Order
from products.models import Product
from django.db.models.functions import TruncMonth
from django.db.models.functions import ExtractYear
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter




@login_required
def sales_report(request):
    # Default to last 30 days
    end_date = timezone.now()
    start_date = end_date - timedelta(days=30)
    
    if request.method == 'POST':
        start_date = datetime.strptime(request.POST.get('start_date'), '%Y-%m-%d')
        end_date = datetime.strptime(request.POST.get('end_date'), '%Y-%m-%d')
    
    orders = Order.objects.filter(
        created_at__range=[start_date, end_date],
        status='completed'
    )
    
    total_sales = orders.aggregate(total=Sum('total_amount'))['total'] or 0
    order_count = orders.count()
    
    context = {
        'start_date': start_date.strftime('%Y-%m-%d'),
        'end_date': end_date.strftime('%Y-%m-%d'),
        'total_sales': total_sales,
        'order_count': order_count,
        'orders': orders,
    }
    return render(request, 'reports/sales.html', context)

@login_required
def daily_sales_report(request):
    today = timezone.now().date()
    orders = Order.objects.filter(
        created_at__date=today,
        status='completed'
    )
    
    total_sales = orders.aggregate(total=Sum('total_amount'))['total'] or 0
    order_count = orders.count()
    
    context = {
        'date': today,
        'total_sales': total_sales,
        'order_count': order_count,
        'orders': orders,
    }
    return render(request, 'reports/daily_sales.html', context)

@login_required
def inventory_report(request):
    products = Product.objects.all().order_by('stock')
    low_stock = products.filter(stock__lt=F('min_stock_level'))
    out_of_stock = products.filter(stock=0)
    
    context = {
        'products': products,
        'low_stock': low_stock,
        'out_of_stock': out_of_stock,
    }
    return render(request, 'reports/inventory.html', context)

@login_required
def export_sales_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="sales_report.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Order Number', 'Customer', 'Date', 'Amount', 'Status'])
    
    orders = Order.objects.all().order_by('-created_at')
    for order in orders:
        writer.writerow([
            order.order_number,
            order.customer_name,
            order.created_at.strftime('%Y-%m-%d'),
            order.total_amount,
            order.get_status_display()
        ])
    
    return response
@login_required
def monthly_sales_report(request):
    # Get current year and month
    now = timezone.now()
    year = request.GET.get('year', now.year)
    
    try:
        year = int(year)
    except ValueError:
        year = now.year
    
    # Get monthly sales data
    monthly_data = (
        Order.objects
        .filter(
            created_at__year=year,
            status='completed'
        )
        .annotate(month=TruncMonth('created_at'))
        .values('month')
        .annotate(
            total_sales=Sum('total_amount'),
            order_count=Count('id')
        )
        .order_by('month')
    )
    
    # Prepare data for chart
    months = []
    sales = []
    orders = []
    
    for month_data in monthly_data:
        months.append(month_data['month'].strftime('%B'))
        sales.append(float(month_data['total_sales']))
        orders.append(month_data['order_count'])
    
    context = {
        'year': year,
        'years': range(now.year - 5, now.year + 1),
        'monthly_data': monthly_data,
        'chart_data': {
            'months': months,
            'sales': sales,
            'orders': orders,
        }
    }
    return render(request, 'reports/monthly_sales.html', context)

@login_required
def yearly_sales_report(request):
    # Get yearly sales data
    yearly_data = (
        Order.objects
        .filter(status='completed')
        .annotate(year=ExtractYear('created_at'))
        .values('year')
        .annotate(
            total_sales=Sum('total_amount'),
            order_count=Count('id')
        )
        .order_by('-year')
    )
    
    # Prepare data for chart
    years = []
    sales = []
    orders = []
    
    for year_data in yearly_data:
        years.append(str(year_data['year']))
        sales.append(float(year_data['total_sales']))
        orders.append(year_data['order_count'])
    
    context = {
        'yearly_data': yearly_data,
        'chart_data': {
            'years': years,
            'sales': sales,
            'orders': orders,
        }
    }
    return render(request, 'reports/yearly_sales.html', context)

@login_required
def low_stock_report(request):
    # Get low stock items (below minimum stock level)
    low_stock_items = Product.objects.filter(
        stock__lt=F('min_stock_level')
    ).order_by('stock')
    
    # Calculate total value of low stock items
    total_value = low_stock_items.aggregate(
        total=Sum(F('stock') * F('price'))
    )
    
    context = {
        'low_stock_items': low_stock_items,
        'total_value': total_value['total'] or 0,
        'title': 'Low Stock Report'
    }
    return render(request, 'reports/inventory_list.html', context)

@login_required
def out_of_stock_report(request):
    # Get out of stock items
    out_of_stock_items = Product.objects.filter(
        stock=0
    ).order_by('name')
    
    context = {
        'low_stock_items': out_of_stock_items,
        'total_value': 0,  # No value for out of stock items
        'title': 'Out of Stock Report'
    }
    return render(request, 'reports/inventory_list.html', context)

@login_required
def export_inventory_csv(request):
    report_type = request.GET.get('type', 'low_stock')
    
    if report_type == 'low_stock':
        products = Product.objects.filter(stock__lt=F('min_stock_level'))
        filename = 'low_stock_report.csv'
    else:
        products = Product.objects.filter(stock=0)
        filename = 'out_of_stock_report.csv'
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    writer.writerow(['Product', 'Category', 'Current Stock', 'Min Stock', 'Price', 'Status'])
    
    for product in products:
        writer.writerow([
            product.name,
            str(product.category),
            product.stock,
            product.min_stock_level,
            product.price,
            'Out of Stock' if product.stock == 0 else 'Low Stock'
        ])
    
    return response

@login_required
def product_performance(request):
    # Get product performance data (sales and revenue)
    products = Product.objects.annotate(
        total_sold=Sum('orderitem__quantity'),
        total_revenue=Sum('orderitem__total')
    ).order_by('-total_revenue')
    
    context = {
        'products': products,
        'title': 'Product Performance Report'
    }
    return render(request, 'reports/product_performance.html', context)

@login_required
def top_selling_products(request):
    # Get top 10 selling products
    top_products = Product.objects.annotate(
        total_sold=Sum('orderitem__quantity')
    ).order_by('-total_sold')[:10]
    
    context = {
        'products': top_products,
        'title': 'Top Selling Products'
    }
    return render(request, 'reports/product_performance.html', context)

@login_required
def export_product_csv(request):
    report_type = request.GET.get('type', 'performance')
    
    if report_type == 'performance':
        products = Product.objects.annotate(
            total_sold=Sum('orderitem__quantity'),
            total_revenue=Sum('orderitem__total')
        ).order_by('-total_revenue')
        filename = 'product_performance.csv'
    else:
        products = Product.objects.annotate(
            total_sold=Sum('orderitem__quantity')
        ).order_by('-total_sold')[:10]
        filename = 'top_selling_products.csv'
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    writer.writerow(['Product', 'Category', 'Units Sold', 'Total Revenue', 'Average Price', 'Current Stock'])
    
    for product in products:
        avg_price = product.total_revenue / product.total_sold if product.total_sold else 0
        writer.writerow([
            product.name,
            str(product.category),
            product.total_sold or 0,
            product.total_revenue or 0,
            round(avg_price, 2),
            product.stock
        ])
    
    return response

@login_required
def low_performing_products(request):
    # Get products with low sales (bottom 10 by revenue)
    low_performers = Product.objects.annotate(
        total_sold=Sum('orderitem__quantity'),
        total_revenue=Sum('orderitem__total')
    ).order_by('total_revenue')[:10]  # Get 10 lowest performers
    
    context = {
        'products': low_performers,
        'title': 'Low Performing Products'
    }
    return render(request, 'reports/product_performance.html', context)

@login_required
def export_product_csv(request):
    report_type = request.GET.get('type', 'performance')
    
    if report_type == 'performance':
        products = Product.objects.annotate(
            total_sold=Sum('orderitem__quantity'),
            total_revenue=Sum('orderitem__total')
        ).order_by('-total_revenue')
        filename = 'product_performance.csv'
    elif report_type == 'top':
        products = Product.objects.annotate(
            total_sold=Sum('orderitem__quantity')
        ).order_by('-total_sold')[:10]
        filename = 'top_selling_products.csv'
    else:  # low performers
        products = Product.objects.annotate(
            total_sold=Sum('orderitem__quantity'),
            total_revenue=Sum('orderitem__total')
        ).order_by('total_revenue')[:10]
        filename = 'low_performing_products.csv'
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    writer.writerow(['Product', 'Category', 'Units Sold', 'Total Revenue', 'Average Price', 'Current Stock'])
    
    for product in products:
        avg_price = product.total_revenue / product.total_sold if product.total_sold else 0
        writer.writerow([
            product.name,
            str(product.category),
            product.total_sold or 0,
            product.total_revenue or 0,
            round(avg_price, 2),
            product.stock
        ])
    
    return response
@login_required
def export_sales_excel(request):
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"
    
    # Add headers with formatting
    headers = ['Order Number', 'Customer', 'Date', 'Amount', 'Status']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = header
        ws[f'{col_letter}1'].font = Font(bold=True)
        ws[f'{col_letter}1'].alignment = Alignment(horizontal='center')
    
    # Get sales data
    orders = Order.objects.all().order_by('-created_at')
    
    # Add data rows
    for row_num, order in enumerate(orders, 2):
        ws[f'A{row_num}'] = order.order_number
        ws[f'B{row_num}'] = order.customer_name
        ws[f'C{row_num}'] = order.created_at.strftime('%Y-%m-%d')
        ws[f'D{row_num}'] = float(order.total_amount)
        ws[f'E{row_num}'] = order.get_status_display()
    
    # Format columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    
    # Format currency column
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            cell.number_format = '$#,##0.00'
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=sales_report.xlsx'
    wb.save(response)
    
    return response